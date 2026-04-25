"""Generate an RTD workbook for BOVA W5 weekly options.

Writes calls (BOVAD) and puts (BOVAP) for integer strikes within +/-10%
of a chosen center strike, using the same Asset / Strike / Cont. Abertos
layout as generate_rtd_excel_from_b3.py.
"""
from __future__ import annotations

import argparse
import math
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

HEADER_FILL = "1F4E78"
DEFAULT_WORKSHEET = "RTD_OI"
DEFAULT_SUFFIX = "B_0"
DEFAULT_SERIES = "W5"
DEFAULT_CENTER = 187
DEFAULT_PCT = 0.10


def _build_rows(center: int, pct: float, series: str, suffix: str):
    low = math.floor(center * (1 - pct))
    high = math.ceil(center * (1 + pct))
    strikes = list(range(low, high + 1))

    rows = []
    for prefix in ("BOVAD", "BOVAP"):
        for k in strikes:
            ticker = f"{prefix}{k}{series}"
            strike_formula = f'=RTD("RTDTrading.RTDServer",,"{ticker}_{suffix}","PEX")'
            oi_formula = f'=RTD("RTDTrading.RTDServer",,"{ticker}_{suffix}","CAB")'
            rows.append((ticker, strike_formula, oi_formula))
    return rows, low, high


def _write_workbook(rows, output_path: Path, worksheet_name: str, center: int, low: int, high: int, series: str):
    wb = Workbook()
    ws = wb.active
    ws.title = worksheet_name

    ws.append(["Asset", "Strike", "Cont. Abertos"])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor=HEADER_FILL)

    for row in rows:
        ws.append(list(row))

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:C{ws.max_row}"
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18

    meta = wb.create_sheet("README")
    meta["A1"] = "Usage"
    meta["A1"].font = Font(bold=True)
    meta["A2"] = f"Series: {series}"
    meta["A3"] = f"Center strike: {center}"
    meta["A4"] = f"Strike range: {low} - {high}"
    meta["A5"] = f"Rows (calls + puts): {len(rows)}"
    meta["A6"] = "1. Open this workbook in Excel with Profit Pro RTD available."
    meta["A7"] = "2. Let the RTD formulas populate (illiquid strikes will show #N/A)."
    meta["A8"] = "3. Run export_rtd_oi_from_excel.bat to snapshot evaluated values."
    meta.column_dimensions["A"].width = 110

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(description="Generate RTD workbook for BOVA W5 weekly options")
    parser.add_argument("--center", type=int, default=DEFAULT_CENTER, help="Center strike (default: 187)")
    parser.add_argument("--pct", type=float, default=DEFAULT_PCT, help="Range fraction around center (default: 0.10)")
    parser.add_argument("--series", default=DEFAULT_SERIES, help="Series suffix appended to strike (default: W5)")
    parser.add_argument("--suffix", default=DEFAULT_SUFFIX, help="Profit RTD suffix (default: B_0)")
    parser.add_argument("--worksheet", default=DEFAULT_WORKSHEET, help="Worksheet name (default: RTD_OI)")
    parser.add_argument("--output", default=None, help="Output workbook path")
    args = parser.parse_args()

    script_dir = Path(__file__).resolve().parent
    output_path = Path(args.output) if args.output else script_dir / f"RTD_OI_BOVA_{args.series}.xlsx"

    rows, low, high = _build_rows(args.center, args.pct, args.series, args.suffix)
    _write_workbook(rows, output_path, args.worksheet, args.center, low, high, args.series)

    print(f"CREATED {output_path}")
    print(f"SERIES {args.series}")
    print(f"CENTER {args.center}")
    print(f"RANGE {low}-{high}")
    print(f"ROWS {len(rows)}")


if __name__ == "__main__":
    main()
