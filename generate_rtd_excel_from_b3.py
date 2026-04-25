from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from get_b3_data import extract_options_oi, fetch_b3_historical_file, search_b3_historical_file


HEADER_FILL = "1F4E78"
DEFAULT_SUFFIX = "B_0"
DEFAULT_WORKSHEET = "RTD_OI"


def _load_options(underlying: str, date: str | None):
    raw = fetch_b3_historical_file(date)
    source_date = date
    if raw.empty:
        print("[*] Primary B3 date returned no data; searching recent business days...")
        raw = search_b3_historical_file(max_attempts=7)
        source_date = "recent_business_day"
    if raw.empty:
        raise RuntimeError("No B3 historical option data found in the last 7 business days")

    options = extract_options_oi(raw, underlying)
    if options.empty:
        raise RuntimeError(f"No options found for {underlying} in the selected B3 data")

    options = options.drop_duplicates(subset=["ticker"]).copy()
    options["type_rank"] = options["type"].map({"call": 0, "put": 1}).fillna(2)
    options = options.sort_values(["type_rank", "strike", "ticker"]).reset_index(drop=True)
    return options, source_date


def _write_workbook(options, output_path: Path, underlying: str, source_date: str | None, worksheet_name: str, suffix: str):
    wb = Workbook()
    ws = wb.active
    ws.title = worksheet_name

    headers = ["Asset", "Strike", "Cont. Abertos"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor=HEADER_FILL)

    for row in options.itertuples(index=False):
        ticker = str(row.ticker).strip().upper()
        strike_formula = f'=RTD("RTDTrading.RTDServer",,"{ticker}_{suffix}","PEX")'
        oi_formula = f'=RTD("RTDTrading.RTDServer",,"{ticker}_{suffix}","CAB")'
        ws.append([ticker, strike_formula, oi_formula])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:C{ws.max_row}"
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18

    meta = wb.create_sheet("README")
    meta["A1"] = "Usage"
    meta["A1"].font = Font(bold=True)
    meta["A2"] = f"Underlying: {underlying.upper()}"
    meta["A3"] = f"B3 source date: {source_date or 'auto'}"
    meta["A4"] = f"Contracts loaded: {len(options)}"
    meta["A5"] = f"Worksheet name: {worksheet_name}"
    meta["A6"] = "1. Open this workbook in Excel with Profit Pro RTD available."
    meta["A7"] = "2. Let the RTD formulas populate."
    meta["A8"] = "3. Run export_rtd_oi_from_excel.bat to export evaluated values to RTD_OI.csv."
    meta.column_dimensions["A"].width = 110

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def _default_output_path(script_dir: Path, underlying: str) -> Path:
    return script_dir / f"RTD_OI_{underlying.upper()}_from_B3.xlsx"


def main():
    parser = argparse.ArgumentParser(description="Generate an RTD Excel workbook from B3 option tickers")
    parser.add_argument("--underlying", default="BOVA11", help="Underlying symbol, e.g. BOVA11")
    parser.add_argument("--date", default=None, help="B3 date in YYYY-MM-DD format; falls back to recent business day if unavailable")
    parser.add_argument("--output", default=None, help="Output workbook path")
    parser.add_argument("--worksheet", default=DEFAULT_WORKSHEET, help="Worksheet name for the RTD sheet")
    parser.add_argument("--suffix", default=DEFAULT_SUFFIX, help="Profit RTD suffix appended to each ticker")
    args = parser.parse_args()

    script_dir = Path(__file__).resolve().parent
    output_path = Path(args.output) if args.output else _default_output_path(script_dir, args.underlying)

    options, source_date = _load_options(args.underlying, args.date)
    _write_workbook(
        options=options,
        output_path=output_path,
        underlying=args.underlying,
        source_date=source_date,
        worksheet_name=args.worksheet,
        suffix=args.suffix,
    )

    print(f"CREATED {output_path}")
    print(f"UNDERLYING {args.underlying.upper()}")
    print(f"CONTRACTS {len(options)}")
    print(f"FIRST {options.iloc[0]['ticker']}")
    print(f"LAST {options.iloc[-1]['ticker']}")
    print(f"WORKSHEET {args.worksheet}")


if __name__ == "__main__":
    main()