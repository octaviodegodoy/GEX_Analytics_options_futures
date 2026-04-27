"""Generate an RTD workbook for option strikes within +/- pct of spot.

Layout matches the existing RTD workbooks (Asset / Strike / Cont. Abertos)
with one row per ticker:

    Asset            Strike                                         Cont. Abertos
    BOVAD187W5       =RTD("RTDTrading.RTDServer";;"BOVAD187W5_B_0";"PEX")
    BOVAP187         =RTD("RTDTrading.RTDServer";;"BOVAP187_B_0";"CAB")

By default it emits both monthly (no series tag) and W5 weekly series for
the current calendar month, for both calls and puts, at every integer strike
within +/- 7% of the supplied spot.

Usage
-----
    python generate_rtd_excel_options.py --spot 187.50
    python generate_rtd_excel_options.py --underlying BOVA --spot 187.5 --pct 0.07
    python generate_rtd_excel_options.py --spot 28 --underlying VALE --month-call F --month-put R
"""
from __future__ import annotations

import argparse
import datetime as _dt
import math
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


HEADER_FILL = "1F4E78"
DEFAULT_WORKSHEET = "RTD_OI"
DEFAULT_SUFFIX = "B_0"
DEFAULT_PCT = 0.07
DEFAULT_UNDERLYING = "BOVA"
DEFAULT_WEEKLY_TAGS = ("W5",)  # add "W1","W2","W3","W4" via CLI if desired

# B3 option-series month letters
_CALL_MONTH_LETTERS = "ABCDEFGHIJKL"   # Jan..Dec
_PUT_MONTH_LETTERS = "MNOPQRSTUVWX"    # Jan..Dec


def _month_letters(month: int) -> tuple[str, str]:
    if not 1 <= month <= 12:
        raise ValueError(f"month must be 1..12, got {month}")
    return _CALL_MONTH_LETTERS[month - 1], _PUT_MONTH_LETTERS[month - 1]


def _build_rows(
    spot: float,
    pct: float,
    underlying: str,
    call_letter: str,
    put_letter: str,
    weekly_tags: tuple[str, ...],
    suffix: str,
):
    if spot <= 0:
        raise ValueError("spot must be positive")
    if pct <= 0:
        raise ValueError("pct must be positive")

    low = math.floor(spot * (1 - pct))
    high = math.ceil(spot * (1 + pct))
    strikes = list(range(int(low), int(high) + 1))

    underlying = underlying.strip().upper()
    # Series tags: "" = monthly; "W1".."W5" = weekly
    series_tags = ("",) + tuple(weekly_tags)

    rows: list[tuple[str, str, str]] = []
    seen: set[str] = set()
    for tag in series_tags:
        for letter in (call_letter, put_letter):
            for k in strikes:
                ticker = f"{underlying}{letter}{k}{tag}".upper()
                if ticker in seen:
                    continue
                seen.add(ticker)
                topic = f"{ticker}_{suffix}"
                strike_formula = f'=RTD("RTDTrading.RTDServer",,"{topic}","PEX")'
                oi_formula = f'=RTD("RTDTrading.RTDServer",,"{topic}","CAB")'
                rows.append((ticker, strike_formula, oi_formula))
    return rows, low, high, strikes


def _write_workbook(
    rows,
    output_path: Path,
    worksheet_name: str,
    underlying: str,
    spot: float,
    pct: float,
    low: int,
    high: int,
    series_summary: str,
):
    wb = Workbook()
    ws = wb.active
    ws.title = worksheet_name

    ws.append(["Asset", "Strike", "Cont. Abertos"])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor=HEADER_FILL)

    for row in rows:
        ws.append(list(row))

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:C{ws.max_row}"
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20

    meta = wb.create_sheet("README")
    meta["A1"] = "Usage"
    meta["A1"].font = Font(bold=True)
    meta["A2"] = f"Underlying: {underlying.upper()}"
    meta["A3"] = f"Spot reference: {spot:.4f}"
    meta["A4"] = f"Range: +/- {pct * 100:.2f}%  ->  strikes {low} .. {high}"
    meta["A5"] = f"Series included: {series_summary}"
    meta["A6"] = f"Rows (calls + puts, all series): {len(rows)}"
    meta["A7"] = "1. Open this workbook in Excel with Profit Pro RTD available."
    meta["A8"] = "2. Let the RTD formulas populate (illiquid strikes will show #N/A)."
    meta["A9"] = "3. Run export_rtd_oi_from_excel.bat to snapshot evaluated values."
    meta.column_dimensions["A"].width = 110

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def _default_output_path(script_dir: Path, underlying: str, pct: float) -> Path:
    pct_tag = f"{int(round(pct * 100))}pct"
    return script_dir / f"RTD_OI_{underlying.upper()}_{pct_tag}.xlsx"


def main():
    today = _dt.date.today()
    default_call, default_put = _month_letters(today.month)

    parser = argparse.ArgumentParser(
        description="Generate RTD workbook for option strikes within +/- pct of spot",
    )
    parser.add_argument("--spot", type=float, required=True, help="Spot price of underlying (e.g. 187.5)")
    parser.add_argument("--pct", type=float, default=DEFAULT_PCT, help="Range fraction around spot (default: 0.07)")
    parser.add_argument("--underlying", default=DEFAULT_UNDERLYING, help="Ticker root, e.g. BOVA, VALE, PETR (default: BOVA)")
    parser.add_argument("--month-call", default=default_call, help=f"Call month letter A..L (default for current month: {default_call})")
    parser.add_argument("--month-put", default=default_put, help=f"Put month letter M..X (default for current month: {default_put})")
    parser.add_argument(
        "--weekly",
        nargs="*",
        default=list(DEFAULT_WEEKLY_TAGS),
        help='Weekly series tags to include (default: W5). Use "" or --weekly with no values to skip weeklies.',
    )
    parser.add_argument("--suffix", default=DEFAULT_SUFFIX, help="Profit RTD market suffix (default: B_0)")
    parser.add_argument("--worksheet", default=DEFAULT_WORKSHEET, help="Worksheet name (default: RTD_OI)")
    parser.add_argument("--output", type=Path, default=None, help="Output .xlsx path (default: RTD_OI_<UND>_<pct>pct.xlsx next to script)")
    args = parser.parse_args()

    script_dir = Path(__file__).resolve().parent
    output_path = args.output or _default_output_path(script_dir, args.underlying, args.pct)

    weekly_tags = tuple(t.strip().upper() for t in (args.weekly or ()) if t and t.strip())

    call_letter = args.month_call.strip().upper()
    put_letter = args.month_put.strip().upper()
    if call_letter not in _CALL_MONTH_LETTERS:
        parser.error(f"--month-call must be one of {_CALL_MONTH_LETTERS}, got {call_letter!r}")
    if put_letter not in _PUT_MONTH_LETTERS:
        parser.error(f"--month-put must be one of {_PUT_MONTH_LETTERS}, got {put_letter!r}")

    rows, low, high, strikes = _build_rows(
        spot=args.spot,
        pct=args.pct,
        underlying=args.underlying,
        call_letter=call_letter,
        put_letter=put_letter,
        weekly_tags=weekly_tags,
        suffix=args.suffix,
    )

    series_summary = "monthly" + ("" if not weekly_tags else f" + weekly[{', '.join(weekly_tags)}]")
    _write_workbook(
        rows=rows,
        output_path=output_path,
        worksheet_name=args.worksheet,
        underlying=args.underlying,
        spot=args.spot,
        pct=args.pct,
        low=low,
        high=high,
        series_summary=series_summary,
    )

    print(f"[OK] Wrote {output_path}")
    print(f"     Underlying : {args.underlying.upper()}")
    print(f"     Spot       : {args.spot:.4f}")
    print(f"     Range      : +/- {args.pct * 100:.2f}%  ->  strikes {low} .. {high} ({len(strikes)} strikes)")
    print(f"     Series     : {series_summary}")
    print(f"     Rows       : {len(rows)} RTD formulas")


if __name__ == "__main__":
    main()
